import openpyxl

class ExcelSaver:

    _fields = ['Channel link', 'Title', 'Subscribers', 'Email']

    def __init__(self):
        pass

    def export(self, result, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        self.define_titles(ws)
        
        i = 2
        for line in result:
            j = 2
            for key in result[line]:
                value = result[line][key]

                if key == 'channelId':
                    value = 'https://www.youtube.com/channel/' + value

                ws.cell(row=i, column=j).value = value
                j += 1
            i += 1

        self.adjust_size(ws)
        wb.save(path + '.xlsx')

    def define_titles(self, ws):
        i = 2
        for f in ExcelSaver._fields:
            ws.cell(row=1, column=i).value = f
            i += 1

    def adjust_size(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
